#!/usr/bin/env python3
"""
Convert CSV to Excel and embed images referenced in the CSV.

Usage: python combine_excel_image.py <folder_path> <output_path>
"""
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path


def combine_excel_image(folder_path, output_path):
    """
    Convert CSV to Excel and embed images referenced in the CSV.
    
    Args:
        folder_path (str): Path to folder containing CSV and image files
        output_path (str): Path where to save the output Excel file (can be directory or file)
        
    Returns:
        bool: True if successful, False if failed
    """
    try:
        folder_path = Path(folder_path)
        output_path = Path(output_path)
        
        # If output_path is a directory or doesn't have .xlsx extension, treat it as directory
        if output_path.is_dir() or not output_path.suffix.lower() == '.xlsx':
            output_path = output_path / 'output.xlsx'
        
        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Find CSV files
        csv_files = list(folder_path.glob("*.csv"))
        if not csv_files:
            print(f"No CSV files found in {folder_path}", file=sys.stderr)
            return False
        
        csv_file = csv_files[0]
        print(f"Reading CSV: {csv_file}", file=sys.stderr)
        
        # Read CSV
        df = pd.read_csv(csv_file)
        
        # Create output Excel file
        df.to_excel(output_path, index=False, sheet_name="Sheet1")
        print(f"Excel created: {output_path}", file=sys.stderr)
        
        # Load workbook to insert images
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Get all image files in the folder
        image_extensions = {".jpg", ".jpeg", ".png", ".gif", ".bmp"}
        image_files = {
            img.stem + img.suffix: img 
            for img in folder_path.iterdir() 
            if img.suffix.lower() in image_extensions
        }
        
        print(f"Found {len(image_files)} images", file=sys.stderr)
        
        # Set default cell dimensions to accommodate images
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 100
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        
        # Iterate through cells and replace image references with actual images
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_value = cell.value
                
                if cell_value and isinstance(cell_value, str):
                    # Check if cell contains image filename
                    image_filename = cell_value.strip()
                    
                    if image_filename in image_files:
                        image_path = image_files[image_filename]
                        
                        try:
                            # Create image and size it to fit inside cell
                            img = XLImage(str(image_path))
                            # Set image dimensions to fit nicely inside the cell
                            img.width = 95
                            img.height = 95
                            
                            # Add image to worksheet at cell position
                            ws.add_image(img, cell.coordinate)
                            
                            # Clear the cell value
                            cell.value = None
                            
                            print(f"Embedded image at {cell.coordinate}: {image_filename}", file=sys.stderr)
                        except Exception as e:
                            print(f"Error embedding image at {cell.coordinate}: {e}", file=sys.stderr)
        
        # Save workbook
        wb.save(output_path)
        print(f"Excel with embedded images saved: {output_path}", file=sys.stderr)
        
        return True
        
    except Exception as e:
        print(f"Error processing Excel: {e}", file=sys.stderr)
        return False


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python combine_excel_image.py <folder_path> <output_path>", file=sys.stderr)
        sys.exit(1)
    
    folder = sys.argv[1]
    output = sys.argv[2]
    
    if combine_excel_image(folder, output):
        sys.exit(0)
    else:
        sys.exit(1)


print("hi")